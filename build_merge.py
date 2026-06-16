# -*- coding: utf-8 -*-
# מיזוג מצבת משרד המשפטים (גיליון "מכשירים פעילים") עם השלמה מקובץ "מתקן מים"
import openpyxl, warnings, datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

DIR = "/root/.claude/uploads/2ef5db8b-6650-546a-9720-1f517b3a7ae0"
F_BASE = f"{DIR}/d7a88770-_______________________.xlsx"   # קובץ 2 - בסיס
F_SRC  = f"{DIR}/7477d277-_______________________.xlsx"   # קובץ 1 - מקור השלמה
OUT    = "/home/user/claude-skills/מצבת_משרד_המשפטים_מאוחדת.xlsx"

def norm(v):
    """נרמול מפתח למחרוזת ללא .0 וללא רווחים"""
    if v is None: return ""
    s = str(v).strip()
    if s.endswith(".0"): s = s[:-2]
    return s

# ---- 1. קריאת מקור ההשלמה (קובץ 1, גיליון "מתקן מים") ----
wb_s = openpyxl.load_workbook(F_SRC, read_only=True, data_only=True)
ws_s = wb_s["מתקן מים"]
rows_s = list(ws_s.iter_rows(values_only=True))
hdr_s = [norm(h) for h in rows_s[0]]
def col_s(name):
    for i, h in enumerate(hdr_s):
        if name in h: return i
    raise KeyError(name)
i_serial = col_s("מספר סידורי")
src_cols = {"דגם": col_s("דגם"), "יחידה": col_s("יחידה"), "אתר": col_s("אתר"),
            "עיר": col_s("עיר"), "מיקום": col_s("מיקום"), "כתובת מלאה": col_s("כתובת מלאה")}

src_all = {}   # serial -> list of records
for r in rows_s[1:]:
    key = norm(r[i_serial])
    if not key: continue
    rec = {k: (r[idx] if idx < len(r) and r[idx] is not None else "") for k, idx in src_cols.items()}
    src_all.setdefault(key, []).append(rec)

def sig(rec):
    """חתימת רשומה לזיהוי כפילות זהה מול סותרת (נרמול רווחים)"""
    return tuple(str(rec[k]).strip() for k in src_cols)

src_map = {}     # serial -> rec נבחר
conflict = set() # serials עם כפילות סותרת
for key, recs in src_all.items():
    # בחירת הרשומה המלאה ביותר
    best = max(recs, key=lambda rc: sum(1 for v in rc.values() if str(v).strip() != ""))
    src_map[key] = best
    if len({sig(rc) for rc in recs}) > 1:   # כפילות סותרת
        conflict.add(key)
wb_s.close()
print(f"כפילויות סותרות במקור: {len(conflict)} -> {sorted(conflict)[:10]}")
print(f"מקור: {len(src_map)} מספרים סידוריים ייחודיים")

# ---- 2. קריאת בסיס (קובץ 2, גיליון "מכשירים פעילים", כותרת שורה 3) ----
wb_b = openpyxl.load_workbook(F_BASE, data_only=True)
ws_b = wb_b["מכשירים פעילים"]
rows_b = list(ws_b.iter_rows(values_only=True))
HDR_ROW = 2  # אינדקס 0-based -> שורה 3
base_hdr = [(h if h is not None else "") for h in rows_b[HDR_ROW]]
ncols = len(base_hdr)
i_tavua = next(i for i, h in enumerate(base_hdr) if "מספר טבוע" in str(h))
base_data = [r for r in rows_b[HDR_ROW+1:] if any(c is not None and str(c).strip() != "" for c in r)]
wb_b.close()
print(f"בסיס: {len(base_data)} שורות מכשירים פעילים")

# עמודות בסיס להסרה (לפי בקשת רון)
DROP = ["מספר תיק לקוח", "מספר מסמך", "תאריך החלפה אחרון - מטהר מים",
        "תאריך החלפה אחרון – נורה", "תאריך החלפה הבא - נורה", "תאריך החלפה הבא - מטהר מים"]
def is_drop(h):
    hs = str(h).strip()
    return any(d.replace("–","-").strip() == hs.replace("–","-") for d in DROP)
keep_idx = [i for i, h in enumerate(base_hdr) if not is_drop(h)]
base_hdr_kept = [base_hdr[i] for i in keep_idx]
print(f"עמודות בסיס: {ncols} -> נשמרות {len(keep_idx)} (הוסרו {ncols-len(keep_idx)})")

# ---- 3. בניית חוברת פלט מעוצבת ----
NEW_COLS = ["דגם", "יחידה", "אתר", "עיר", "מיקום (קובץ מתקנים)", "כתובת מלאה (קובץ מתקנים)", "סטטוס התאמה"]
SRC_KEYS = ["דגם", "יחידה", "אתר", "עיר", "מיקום", "כתובת מלאה"]
# מבנה פלט: # | עמודות בסיס שנשמרו | 6 מושלמות | סטטוס
full_hdr = ["#"] + list(base_hdr_kept) + NEW_COLS
total_cols = len(full_hdr)
nkept = len(base_hdr_kept)          # מס' עמודות בסיס שנשמרו
src_start = 1 + nkept              # אינדקס 0-based של תחילת העמודות המושלמות (אחרי #)
status_pos0 = total_cols - 1       # אינדקס 0-based של עמודת הסטטוס

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "מצבת מאוחדת"
ws.sheet_view.rightToLeft = True

# צבעים וסגנונות
NAVY   = PatternFill("solid", fgColor="1F3864")
ADDED  = PatternFill("solid", fgColor="DDEBF7")  # תכלת בהיר לעמודות שהושלמו
ZEBRA  = PatternFill("solid", fgColor="F2F6FC")
GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
ORANGE = PatternFill("solid", fgColor="FFEB9C")
TITLEF = PatternFill("solid", fgColor="0F2147")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HFONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CFONT = Font(name="Calibri", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center", wrap_text=True)

# שורת כותרת-על ממוזגת
today = datetime.date.today().strftime("%d/%m/%Y")
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
t = ws.cell(row=1, column=1, value=f"משרד המשפטים  |  מצבת מתקני שתייה מאוחדת  |  הופק: {today}")
t.fill = TITLEF; t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# שורת כותרות עמודות (שורה 2)
HROW = 2
for c, name in enumerate(full_hdr, start=1):
    cell = ws.cell(row=HROW, column=c, value=name)
    cell.fill = NAVY; cell.font = HFONT; cell.alignment = CENTER; cell.border = BORDER
ws.row_dimensions[HROW].height = 38

# נתונים
n_done = n_miss = n_conf = 0
miss_list = []
date_fmt = "dd/mm/yyyy"
for ridx, r in enumerate(base_data):
    excel_row = HROW + 1 + ridx
    key = norm(r[i_tavua])
    rec = src_map.get(key)
    is_conf = key in conflict
    if rec:
        n_done += 1
        if is_conf: n_conf += 1
    else:
        n_miss += 1
        miss_list.append(key)
    status_val = ("כפילות במקור - לאימות" if (rec and is_conf)
                  else "הושלם" if rec else "לא נמצא במצבת המתקנים")
    for c in range(total_cols):
        if c == 0:
            val = ridx + 1                          # מספור רץ
        elif c < src_start:
            val = r[keep_idx[c-1]] if keep_idx[c-1] < len(r) else None
        elif c < src_start + len(SRC_KEYS):
            val = rec[SRC_KEYS[c - src_start]] if rec else ""
        else:                                       # סטטוס התאמה
            val = status_val
        cell = ws.cell(row=excel_row, column=c+1, value=val)
        cell.font = CFONT; cell.alignment = (CENTER if c == 0 else RIGHT); cell.border = BORDER
        if isinstance(val, datetime.datetime):
            cell.number_format = date_fmt
        # רקע
        if src_start <= c < src_start + len(SRC_KEYS):
            cell.fill = ADDED
        elif ridx % 2 == 1:
            cell.fill = ZEBRA
    # צביעת מצב (לפי מיקומה בעמודות שנשמרו)
    if "מצב" in [str(h).strip() for h in base_hdr_kept]:
        pos = [str(h).strip() for h in base_hdr_kept].index("מצב") + 1  # +1 בגלל עמודת #
        mcell = ws.cell(row=excel_row, column=pos+1)
        if str(mcell.value).strip() == "מותקן": mcell.fill = GREEN
        elif str(mcell.value).strip(): mcell.fill = RED
    # צביעת סטטוס התאמה
    scell = ws.cell(row=excel_row, column=total_cols)
    scell.fill = ORANGE if (rec and is_conf) else GREEN if rec else RED
    scell.alignment = CENTER

# הקפאה, פילטר
ws.freeze_panes = f"A{HROW+1}"
ws.auto_filter.ref = f"A{HROW}:{get_column_letter(total_cols)}{HROW+len(base_data)}"

# רוחב עמודות אוטומטי (לפי תוכן, מוגבל)
for c in range(1, total_cols+1):
    letter = get_column_letter(c)
    maxlen = len(str(full_hdr[c-1]))
    for ridx in range(len(base_data)):
        v = ws.cell(row=HROW+1+ridx, column=c).value
        if v is not None:
            maxlen = max(maxlen, len(str(v)))
    ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 45)

# ---- 4. דף סיכום מנהלים ----
from openpyxl.chart import BarChart, PieChart, Reference
from collections import Counter

i_matzav = next(i for i, h in enumerate(base_hdr) if str(h).strip() == "מצב")
i_addr   = next(i for i, h in enumerate(base_hdr) if "כתובת התקנה" in str(h))
by_city = Counter(); by_state = Counter()
for r in base_data:
    addr = r[i_addr] if i_addr < len(r) else ""
    city = str(addr).split(",")[0].strip() if addr else "ללא כתובת"
    by_city[city] += 1
    st = str(r[i_matzav]).strip() if (i_matzav < len(r) and r[i_matzav] is not None) else "לא צוין"
    by_state[st or "לא צוין"] += 1

su = wb.create_sheet("סיכום מנהלים", 0)   # ראשון בחוברת
su.sheet_view.rightToLeft = True
TITLEF2 = PatternFill("solid", fgColor="0F2147")
KPIF    = PatternFill("solid", fgColor="1F3864")
LBL     = PatternFill("solid", fgColor="DDEBF7")

su.merge_cells("A1:F1")
c = su["A1"]; c.value = f"משרד המשפטים  |  סיכום מנהלים — מצבת מתקני שתייה  |  {today}"
c.fill = TITLEF2; c.font = Font(bold=True, color="FFFFFF", size=16); c.alignment = Alignment("center", vertical="center")
su.row_dimensions[1].height = 34

# כרטיסי KPI
kpis = [("סה\"כ מכשירים פעילים", len(base_data)),
        ("הושלמו ממצבת המתקנים", n_done - n_conf),
        ("כפילות לאימות", n_conf),
        ("לא נמצאו במצבת", n_miss),
        ("מותקנים", by_state.get("מותקן", 0)),
        ("לא מותקנים", sum(v for k,v in by_state.items() if k!="מותקן"))]
col = 1
for label, val in kpis:
    lc = su.cell(row=3, column=col, value=label)
    lc.fill = LBL; lc.font = Font(bold=True, size=10); lc.alignment = CENTER; lc.border = BORDER
    vc = su.cell(row=4, column=col, value=val)
    vc.fill = KPIF; vc.font = Font(bold=True, color="FFFFFF", size=20); vc.alignment = CENTER; vc.border = BORDER
    su.column_dimensions[get_column_letter(col)].width = 18
    col += 1
su.row_dimensions[4].height = 36

# טבלת פילוח לפי עיר
r0 = 7
su.cell(row=r0, column=1, value="פילוח לפי עיר").font = Font(bold=True, size=13)
hr = r0 + 1
for j, t in enumerate(["עיר", "מספר מכשירים"]):
    cc = su.cell(row=hr, column=1+j, value=t); cc.fill = NAVY; cc.font = HFONT; cc.alignment = CENTER; cc.border = BORDER
city_sorted = by_city.most_common()
for i, (city, n) in enumerate(city_sorted, start=hr+1):
    a = su.cell(row=i, column=1, value=city); a.border = BORDER; a.alignment = RIGHT
    b = su.cell(row=i, column=2, value=n); b.border = BORDER; b.alignment = CENTER
    if (i-hr) % 2 == 0: a.fill = ZEBRA; b.fill = ZEBRA
city_end = hr + len(city_sorted)

# טבלת פילוח לפי מצב (בצד)
sr = r0 + 1
su.cell(row=r0, column=4, value="פילוח לפי מצב").font = Font(bold=True, size=13)
for j, t in enumerate(["מצב", "מספר"]):
    cc = su.cell(row=sr, column=4+j, value=t); cc.fill = NAVY; cc.font = HFONT; cc.alignment = CENTER; cc.border = BORDER
for i, (st, n) in enumerate(by_state.most_common(), start=sr+1):
    a = su.cell(row=i, column=4, value=st); a.border = BORDER; a.alignment = RIGHT
    b = su.cell(row=i, column=5, value=n); b.border = BORDER; b.alignment = CENTER
state_end = sr + len(by_state)

# גרף עמודות - ערים
bar = BarChart(); bar.title = "מכשירים פעילים לפי עיר"; bar.type = "bar"; bar.height = 9; bar.width = 18
data = Reference(su, min_col=2, min_row=hr, max_row=city_end)
cats = Reference(su, min_col=1, min_row=hr+1, max_row=city_end)
bar.add_data(data, titles_from_data=True); bar.set_categories(cats); bar.legend = None
su.add_chart(bar, "G7")

# גרף עוגה - מצב
pie = PieChart(); pie.title = "התפלגות מצב"; pie.height = 8; pie.width = 10
pdata = Reference(su, min_col=5, min_row=sr, max_row=state_end)
pcats = Reference(su, min_col=4, min_row=sr+1, max_row=state_end)
pie.add_data(pdata, titles_from_data=True); pie.set_categories(pcats)
su.add_chart(pie, "G28")

for L in ["D", "E"]:
    su.column_dimensions[L].width = 16
su.column_dimensions["A"].width = 22

wb.save(OUT)
print(f"נשמר: {OUT}")
print(f"ערים בפילוח: {len(city_sorted)} | מצבים: {dict(by_state)}")
print(f"הושלמו: {n_done} (מתוכם כפילות סותרת לאימות: {n_conf}) | לא נמצאו: {n_miss}")
print("דוגמת לא-נמצאו:", miss_list[:10])
