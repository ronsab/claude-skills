from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "טופס הוצאות והכנסות"
ws.sheet_view.rightToLeft = True

header_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
expense_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
income_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
total_font = Font(name="Arial", size=12, bold=True)
total_fill = PatternFill(start_color="F4D03F", end_color="F4D03F", fill_type="solid")
center = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.cell(row=1, column=1, value="הוצאות").font = header_font
ws.cell(row=1, column=1).fill = expense_fill
ws.cell(row=1, column=1).alignment = center

ws.cell(row=1, column=2, value="הכנסות").font = header_font
ws.cell(row=1, column=2).fill = income_fill
ws.cell(row=1, column=2).alignment = center

rows_count = 20
for row in range(2, rows_count + 2):
    for col in range(1, 3):
        cell = ws.cell(row=row, column=col)
        cell.alignment = center
        cell.border = border
        cell.number_format = '#,##0.00 ₪'

ws.cell(row=1, column=1).border = border
ws.cell(row=1, column=2).border = border

total_row = rows_count + 2
ws.cell(row=total_row, column=1, value=f"=SUM(A2:A{rows_count + 1})").font = total_font
ws.cell(row=total_row, column=1).fill = total_fill
ws.cell(row=total_row, column=1).alignment = center
ws.cell(row=total_row, column=1).border = border
ws.cell(row=total_row, column=1).number_format = '#,##0.00 ₪'

ws.cell(row=total_row, column=2, value=f"=SUM(B2:B{rows_count + 1})").font = total_font
ws.cell(row=total_row, column=2).fill = total_fill
ws.cell(row=total_row, column=2).alignment = center
ws.cell(row=total_row, column=2).border = border
ws.cell(row=total_row, column=2).number_format = '#,##0.00 ₪'

balance_row = total_row + 1
ws.cell(row=balance_row, column=1, value="יתרה:").font = total_font
ws.cell(row=balance_row, column=1).alignment = center
ws.cell(row=balance_row, column=1).border = border

ws.cell(row=balance_row, column=2, value=f"=B{total_row}-A{total_row}").font = total_font
ws.cell(row=balance_row, column=2).alignment = center
ws.cell(row=balance_row, column=2).border = border
ws.cell(row=balance_row, column=2).number_format = '#,##0.00 ₪'

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 25
ws.row_dimensions[1].height = 30

wb.save("/home/user/claude-skills/טופס_הוצאות_והכנסות.xlsx")
print("Created: טופס_הוצאות_והכנסות.xlsx")
